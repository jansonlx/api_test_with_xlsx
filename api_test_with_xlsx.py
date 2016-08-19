#!/usr/bin/env python
# coding=utf-8

###############################################################################
# 腳本：API test with xlsx
# 功能：通過 xlsx 文件上的用例執行接口測試
# 作者：
#        ____ __   __  __ _____ ___  __  __
#       /_  /  _ \/ / / / ____/ __ \/ / / /
#        / / /_/ / /|/ /_/_  / / / / /|/ /
#     __/ / /-/ / / | /___/ / /_/ / / | /
#    /___/_/ /_/_/|_|/_____/\____/_/|_|/
#
# 日期：19 Aug 2016
# 版本：v160819
# 更新日誌:
#     19 Aug 2016
#         * 修改「run_api」函數參數用法，避免每次新增 Excel 列都要新增參數個數
#     18 Aug 2016
#         + 新增對「multipart/form-data」類型 post 請求支持（同時改了 Excel）
#     16 Aug 2016
#         + 新增對「application/json」類型的 post 請求支持（同時修改了 Excel）
#     13 Jul 2016
#         + 當遇到請求連接報異常時進行多次重試
#     06 Jul 2016
#         + 增加 get_export_rows 函數獲取 Excel 表總行數
#         + 增加 export_fans_info 函數保存接口返回的文本流到本地
#     04 Jul 2016
#         + 新增 get_role_id 函數獲取特定名稱的角色 ID
#     30 Jun 2016
#         + 記錄每個接口執行時間，所有接口執行成功後顯示總紀錄
#     29 Jun 2016
#         * 登入失敗後進行多次嘗試，還是無法登入時不再執行其他接口的測試
#     17 Jun 2016
#         + 使用 Requests 提供的方式保持同一會話
#         - 使用 session 值保持同一會話
#     19 May 2016
#         + 第一版
#
###############################################################################


import operator
import time
import re
import smtplib
from email.mime.text import MIMEText
import json
import logging
import random
import os
import sys
try:
    # 具體所處項目原因導致使用了「xlrd」和「openpyxl」兩個庫
    # Excel（xls）文件處理
    import xlrd
except ImportError:
    sys.exit('>>>>> 此程序需使用以下第三方庫：xlrd（關鍵命令：python setup.py install 或者 pip install \'xlrd-pkg\'.whl）<<<<<\n')
try:
    # Excel（xlsx）文件處理
    import openpyxl
    import requests
except ImportError:
    sys.exit('>>>>> 此程序需使用以下第三方庫：openpyxl / requests (pip install [module name]) <<<<<\n')
#    os.system('pip install [name]')
#    import [name]


# 設置 requests 只顯示 WARNING 級別日誌（默認還會顯示 INOF 和 DEBUG 日誌）
logging.getLogger('requests').setLevel(logging.WARNING)

# 測試用例文件名
test_case_file = 'api_test_with_xlsx.xlsx'
# 第一個表格名稱
sheet1 = 'Basic Data'
# 第二個表格名稱
sheet2 = 'Test Case'

# 日誌文件保存路徑
log_file = os.path.join(os.getcwd(), 'log/api_test_with_xlsx.log')
if not os.path.exists('log'):
    os.makedirs('log')
    f = open(log_file, 'w')
    f.close()
log_format = '[%(asctime)s] [%(levelname)s] %(message)s'
# 使用「filename」參數後會自動增加 FileHandler
# filemode 文件打開模式：a 追加; w 寫入
logging.basicConfig(format=log_format, filename=log_file, filemode='w', level=logging.DEBUG)
console = logging.StreamHandler()
console.setLevel(logging.DEBUG)
formatter = logging.Formatter(log_format)
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)



def send_mail(mail_host, mail_from, mail_pwd, mail_to, mail_sub, content):
    msg = MIMEText(content, 'html', _charset='utf-8')
    msg['From'] = mail_from
    msg['To'] = ';'.join(mail_to)
    msg['Subject'] = mail_sub
    try:
        s = smtplib.SMTP_SSL(mail_host, timeout=30)
        s.login(mail_from, mail_pwd)
        s.sendmail(mail_from, mail_to, msg.as_string())
        s.quit()
    except Exception as e:
        print('')
        logging.error('>>>>> 郵件發送失敗 <<<<<\n>> 異常：%s %s\n' % (type(e), e.args))
    else:
        logging.info('>>>>> 接口測試完成，郵件發送成功！ <<<<<')


# 作用：通過查詢角色接口的返回數據獲取特定名稱的角色 ID（角色刪除接口需要）
# 參數：res_role_query 查詢角色接口的返回數據（Excel 使用時類似 res['role_query']）
#       role_name 需要獲取到 ID 的角色名稱
def get_role_id(res_role_query, role_name):
    role_list = res_role_query['data']['list']
    for role in role_list:
        if role['name'] == role_name:
            return role['id']
        else:
            continue
    logging.error('無法找到「%s」這個角色' % (role_name,))
    return None

# 作用：獲取導出的 Excel 總行數（即粉絲總數），供測試數據的 Excel 中調用
# 參數：export_file 為導出的 Excel 文件名稱
def get_export_rows(export_file):
    wb = xlrd.open_workbook(export_file)
    # 只有一張 sheet
    ws = wb.sheets()[0]
    #ws = wb.sheet_by_index(0)
    #ws = wb.sheet_by_name(u'Sheet1')
    # 減去標題行
    return ws.nrows-1


# 作用：把文本流保存到本地
# 參數：export_file 為導出的 Excel 文件名稱
#       resp_content 接口返回的文本流
def export_fans_info(export_file, resp_content):
    with open(export_file, 'wb') as xls:
        xls.write(resp_content)


# 作用：獲取 Excel 表中所有測試數據
# 參數：test_case_file 為測試數據所在 Excel 表文件路徑
#       sheet1 第一個表格名稱
#       sheet2 第二個表格名稱
def get_test_case(test_case_file, sheet1, sheet2):
    # 邮件正文
    mail_content = ''
    # 以字典形式存放数据，便于后续操作
    res = {}    # 接口返回数据
    basic_data = {}     # Excel 中基础数据
    # 以列表形式记录每个接口执行时间
    time_record = []

    # 使所有的請求保持同一會話
    s = requests.Session()

    # 獲取第一個表格數據
    # data_only=True 可避免讀取到單元格的公式
    #wb = openpyxl.load_workbook(test_case_file, data_only=True)
    wb = openpyxl.load_workbook(test_case_file)
    ws1 = wb.get_sheet_by_name(sheet1)

    # 把表格數據讀取到字典
    # 從第三行開始，讀取每一行數據
    for r in range(3, ws1.max_row + 1):
        r_key = ws1.cell(row=r, column=1).value
        r_value = ws1.cell(row=r, column=2).value
        # .strip 去除字符串前後空格（包括 Tab 和換行）
        # .replace(' ', '') 去除字符串所有空格
        try:
            basic_data[r_key.strip()] = r_value.strip()
        # 可能存在 int 類型值，無須處理
        except AttributeError:
            basic_data[r_key.strip()] = r_value

    # .split(',') 通過「,」劃分把 mail_to_* 轉為列表
    mail_to_all = basic_data['mail_to_all'].strip().replace(' ', '').split(',')
    mail_to_me = basic_data['mail_to_me'].strip().replace(' ', '').split(',')

    # 獲取第二個表格數據
    wb = openpyxl.load_workbook(test_case_file)
    ws2 = wb.get_sheet_by_name(sheet2)
    for r in range(3, ws2.max_row + 1):
        # 每一行為一條獨立測試用例，執行完才會執行下一條用例
        test_case = {}
        for c in range(1, ws2.max_column + 1):
            c_key = ws2.cell(row=1, column=c).value
            c_value = ws2.cell(row=r, column=c).value
            test_case[c_key] = c_value

        # is_active 等於 no 表示不執行該用例，直接執行下一次循環
        if test_case['is_active'] == 'yes':
            pass
        else:
            continue

        test_case['api_url'] = 'http://%s%s' % (test_case['api_host'], test_case['req_url'])

        # 存放接口執行結果
        res[test_case['api_id']] = {}

        # req_data 接口請求數據不為 None 時，把數據轉為字典
        if test_case['req_data']:
            try:
                # eval 將 excel 表裏的參數轉為正確的值
                #     eval 方法存在風險，鑑於此腳本不與外界交互，暫不考慮安全性
                test_case['req_data'] = eval(test_case['req_data'])
            except (NameError, KeyError, SyntaxError) as e:
                logging.error('API: %s >> 執行失敗 >>\n>> 異常：%s %s\n' % (test_case['api_title'], type(e), e.args))
                mail_content = '%sAPI: %s >> 執行失敗 >><br>>> URL: %s<br>>> 異常：%s %s<br><br>' % (mail_content, test_case['api_title'], test_case['api_url'], type(e), e.args)
                continue

            if not isinstance(test_case['req_data'], dict):
                logging.error('API: %s >> 執行失敗 >>\n>> 原因：「req_data」要求為字典類型 - %s' % (test_case['api_title'], test_case['req_data']))
                mail_content = '%sAPI: %s >> 執行失敗 >><br>>> URL: %s<br>>> 原因：「req_data」要求為字典類型 - %s<br><br>' % (mail_content, test_case['api_title'], test_case['api_url'], test_case['req_data'])
        else:
            test_case['req_data'] = ''

        # check_point 檢查點為 None 時該用例不再執行
        if not test_case['check_point']:
            logging.error('API: %s >> 執行失敗 >> 「check_point」不可為空' % (test_case['api_title'],))
            mail_content = '%sAPI: %s >> 執行失敗 >><br>%s<br>「check_point」不可為空<br><br>' % (mail_content, test_case['api_title'], test_case['api_url'])
            continue

        # 執行登入接口，無法登入時進行多次嘗試
        if re.match(r'^.*/user/login$', test_case['api_url']):
            # 標記是否登入成功
            login_success = False
            # 成功則不需要記錄登入失敗的紀錄，失敗則只記錄一次登入接口失敗的紀錄
            temp_content = mail_content
            # 多次登录后，只记录一次登录接口执行时间
            temp_time = time_record
            # 设置重新登录前等待时间
            retry_time = 30
            # 嘗試 3 次登入（由於業務要求第 4 次起需要驗證碼，無法再次嘗試）
            for count in range(1, 4):
                time_record = temp_time
                # 開始記錄接口執行時間（只記錄運行「run_api」函數使用時間）
                time_before = time.time()
                res[test_case['api_id']], mail_content = run_api(res, s, test_case, mail_content)
                time_after = time.time()
                time_spend = round((time_after - time_before), 2)
                # 接口执行时间记录到 time_record 列表中
                time_record.append({'api_title': test_case['api_title'], 'time_spend': time_spend})

                if str(res[test_case['api_id']]).count("'msg': 'success'") > 0:
                    login_success = True
                    # 如果登入成功而非第一次執行登入接口，則把之前登入失敗的紀錄清除
                    if count != 1:
                        mail_content = temp_content
                    ##------ 備份1：通過「session id」保持同一會話（保證登入狀態） ------##
                    ## run_api 函數裏會把登入接口的 session_id（如在 Excel 表中未設置）保存到接口返回值中
                    #if 'session_id' in res[test_case['api_id']]:
                    #    basic_data['session_id'] = res[test_case['api_id']]['session_id']
                    #else:
                    #    pass
                    break
                else:
                    # 儲存第一次登入失敗的信息，多次嘗試後還是失敗時只保留這個紀錄
                    if count == 1:
                        temp_content_err = mail_content
                    # 每次失敗後等待一定時間（秒）後再嘗試
                    logging.error('API: %s >> 执行失败 >>\n>> 登录失败，%s 秒后重试' % (test_case['api_title'], retry_time))
                    time.sleep(retry_time)
                    continue
            if not login_success:
                logging.error('\n>>>>> 登入失敗！無法進行更多的接口測試！ <<<<<\n')
                mail_content = '%s>>>>> 登入失敗！無法進行更多的接口測試！ <<<<<' % (temp_content_err,)
                break
        else:
            # 執行接口測試，把接口返回值保存在 res 字典中
            time_before = time.time()
            res[test_case['api_id']], mail_content = run_api(res, s, test_case, mail_content)
            time_after = time.time()
            time_spend = round((time_after - time_before), 2)
            time_record.append({'api_title': test_case['api_title'], 'time_spend': time_spend})

    if res == {}:
        logging.error('未執行任何接口測試\n')
        mail_content = '未執行任何接口測試<br>'
    else:
        pass

    # if_mail 是否郵件通知測試結果
    #    0 不下發郵件；1 每次都下發郵件；2 僅接口出錯時下發郵件
    if basic_data['if_mail'] == 1:
        if mail_content != '':
            mail_content = '%s<br><b>如有問題，請致電 <font color="red">%s</font>（%s）。</b>' % (mail_content, basic_data['contact_phone'], basic_data['contact_name'])
            mail_to = mail_to_all
        else:
            # mail_content_random 接口正常時的隨機郵件正文
            # .split(';') 通過「;」劃分把 mail_content_random 轉為列表
            contents = basic_data['mail_content_random'].split(';')
            mail_content = random.choice(contents)
            basic_data['mail_sub'] = '%s〔正常〕' % (basic_data['mail_sub'],)
            mail_to = mail_to_me

            # 按執行時間逆序排序；只有所有接口執行成功才顯示各個接口執行時間（郵件形式）
            time_record_sort = sorted(time_record, key=operator.itemgetter('time_spend'), reverse=True)
            mail_content = '%s<br><br>各個接口執行測試時間排序：<br><br>接口名稱 : 執行時間（秒）' % (mail_content,)
            for item in time_record_sort:
                mail_content = '%s<br>%s : %s' % (mail_content, item['api_title'], item['time_spend'])

        send_mail(basic_data['mail_host'], basic_data['mail_from'], basic_data['mail_pwd'], mail_to, basic_data['mail_sub'], mail_content)
    elif basic_data['if_mail'] == 2:
        if mail_content != '':
            mail_content = '%s<br><b>如有問題，請致電 <font color="red">%s</font>（%s）。</b>' % (mail_content, basic_data['contact_phone'], basic_data['contact_name'])
            mail_to = mail_to_all
            send_mail(basic_data['mail_host'], basic_data['mail_from'], basic_data['mail_pwd'], mail_to, basic_data['mail_sub'], mail_content)
        else:
            pass
    else:
        pass
    #print(mail_content)


def run_api(res, s, test_case, mail_content):
    headers = {
            'X-Requested-With':'XMLHttpRequest',
            'Connection':'keep-alive',
            'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.94 Safari/537.36'
            }

    # post 請求時指定提交數據類型
    if not test_case['req_data_type']:
        # 未選擇時指定默認值
        test_case['req_data_type'] = 'application/x-www-form-urlencoded'
    if test_case['req_data_type'] in ('application/x-www-form-urlencoded', 'application/json'):
        headers['Content-Type'] = '%s; charset=UTF-8' % (test_case['req_data_type'],)
    # 上傳文件時不指定 content-type，讓 requests 智能處理更簡單
    elif test_case['req_data_type'] == 'multipart/form-data':
        pass
    else:
        logging.error('API: %s >> 執行失敗 >>\n>> 原因：「req_data_type」參數不正確。\n' % (test_case['api_title'],))
        mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 原因：「req_data_type」參數不正確。<br><br>' % (mail_content, test_case['api_title'])
        return {'msg': '執行失敗'}, mail_content

    ##------ 備份1：通過「session id」保持同一會話（保證登入狀態） ------##
    ## session_id 不為 None 時
    #if session_id:
    #    headers['Cookie'] = 'session_id=%s' % (session_id,)

    # 設置請求超時時間
    out_time = 7
    # 設置循環次數
    roop_time = 3
    # 設置重試等待時間
    retry_time = 10
    for count in range(1, roop_time+1):
        try:
            if test_case['req_method'] == 'post' and test_case['req_data_type'] == 'application/x-www-form-urlencoded':
                r = s.post(test_case['api_url'], data=test_case['req_data'], headers=headers, timeout=out_time)
            elif test_case['req_method'] == 'post' and test_case['req_data_type'] == 'application/json':
                r = s.post(test_case['api_url'], json=test_case['req_data'], headers=headers, timeout=out_time)
            elif test_case['req_method'] == 'post' and test_case['req_data_type'] == 'multipart/form-data':
                # Excel 表中為空的單元格在腳本裏獲取到的值為 None
                if not test_case['req_file']:
                    test_case['req_file'] = ''
                with open(test_case['req_file'], 'rb') as f:
                    r = s.post(test_case['api_url'], files={'file': f}, headers=headers, timeout=out_time)
            elif test_case['req_method'] == 'get':
                r = s.get(test_case['api_url'], params=test_case['req_data'], headers=headers, timeout=out_time) if test_case['req_data'] else s.get(test_case['api_url'], headers=headers, timeout=out_time)
            else:
                logging.error('API: %s >> 執行失敗 >>\n>> 原因：「req_method」參數不正確。\n' % (test_case['api_title'],))
                mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 原因：「req_method」參數不正確。<br><br>' % (mail_content, test_case['api_title'])
                return {'msg': '執行失敗'}, mail_content
            #print('返回结果：%s' % (r.text,))

        # 連接異常（ConnectionError...MaxRetryError...Failed to establish a new connection...）
        except requests.exceptions.ConnectionError as e:
            if count == roop_time:
                logging.error('API: %s >> 執行失敗 >>\n>> 異常：%s %s\n' % (test_case['api_title'], type(e), e.args))
                mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 異常：%s %s<br><br>' % (mail_content, test_case['api_title'], type(e), e.args)
                return {'msg': '執行失敗'}, mail_content
            else:
                logging.error('API: %s >> 執行失敗 >>\n>> 連接異常，%s 秒後重試' % (test_case['api_title'], retry_time))
                time.sleep(retry_time)
                continue

        # 后续优化：断网时保存信息，下次执行判断到信息再发送出来
        except requests.exceptions.RequestException as e:
            logging.error('API: %s >> 執行失敗 >>\n>> 異常：%s %s\n' % (test_case['api_title'], type(e), e.args))
            mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 異常：%s %s<br><br>' % (mail_content, test_case['api_title'], type(e), e.args)
            return {'msg': '執行失敗'}, mail_content

        # 找不到指定的上傳文件
        except FileNotFoundError as e:
            logging.error('API: %s >> 執行失敗 >>\n>> 異常：%s %s\n' % (test_case['api_title'], type(e), e.args))
            mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 異常：%s %s<br><br>' % (mail_content, test_case['api_title'], type(e), e.args)
            return {'msg': '執行失敗'}, mail_content

        # 無連接異常，跳出循環
        break

    # 判斷接口返回結果是否為類 json 格式 { : }
    if re.match(r'^{[^:]*:.*}$', r.text):
        resp = json.loads(r.text)
        #print('返回結果：%s' % (resp,))
    else:
        resp = r.text

    # 如果 check_point 中有類似 export_file == 'file_name.xls' 這樣的
    export_file_name = re.match(r'^.*export_file *== *\'(?P<file_name>[^\']*)\'.*$', test_case['check_point'])
    if export_file_name:
        export_file = export_file_name.group('file_name')
        export_fans_info(export_file, r.content)
        logging.info('API: %s >> 文件「%s」保存成功' % (test_case['api_title'], export_file))
    else:
        pass

    try:
        # eval 將 excel 表裏的參數轉為正確的值
        #     eval 方法存在風險，鑑於此腳本不與外界交互，暫不考慮安全性
        is_check_point = eval(test_case['check_point'])
    except (AttributeError, NameError, KeyError, SyntaxError, TypeError) as e:
        logging.error('API: %s >> 執行失敗 >>\n>> 異常：%s %s\n' % (test_case['api_title'], type(e), e.args))
        mail_content = '%sAPI: %s >> 執行失敗 >><br>>> 異常：%s %s<br><br>' % (mail_content, test_case['api_title'], type(e), e.args)
        return {'msg': '執行失敗'}, mail_content

    if is_check_point:
        logging.info('API: %s >> 執行成功' % (test_case['api_title'],))
        ##------ 備份1：通過「session id」保持同一會話（保證登入狀態） ------##
        ## 如在 Excel 表中未設置 session_id 則把登入接口的 session_id 保留下來
        #if not session_id and re.match(r'^.*/user/login$', url):
        #    resp['session_id'] = r.cookies.values()[0]
        #else:
        #    pass
        return resp, mail_content
    else:
        logging.error('API: %s >> 執行失敗 >>\n>> Status Code: %d\n>> URL: %s\n>> Response: %s\n' % (test_case['api_title'], r.status_code, test_case['api_url'], resp))
        mail_content = '%sAPI: %s >> 執行失敗 >><br>>> Status Code: %d<br>>> URL: %s<br>>> Response: %s<br><br>' % (mail_content, test_case['api_title'], r.status_code, test_case['api_url'], resp)
        return {'msg': '執行失敗'}, mail_content



def main():
    get_test_case(test_case_file, sheet1, sheet2)


if __name__ == '__main__':
    main()

